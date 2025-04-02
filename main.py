# Authors: Conrad Bradford, Blake Rogers, Haley Sommer, Elise Chapman, Rebecca Mecham
# IS 303 Section 003
# Write a program that takes in grade data excel sheets and formats it easy to use.

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


old_workbook =  load_workbook(filename = "Poorly_Organized_Data_1.xlsx")

formatted_workbook = Workbook()

formatted_workbook.remove(formatted_workbook.active)

# COMPLETED AS A GROUP - TASK 1 - Create a new worksheet for each class type (algebra, calc, etc.) 
# This should be dynamic and create the classes based on the student data
list_classes = []
col_num = 1
class_name = ''
for row in range(2, old_workbook['Grades'].max_row + 1) :
    
    new_name = old_workbook['Grades'].cell(row=row, column=col_num).value

    if class_name != new_name :
        list_classes.append(new_name)
        class_name = new_name

for classes in list_classes :
    formatted_workbook.create_sheet(classes)
    formatted_workbook[classes]["A1"] = "Last Name"
    formatted_workbook[classes]["B1"] = "First Name"
    formatted_workbook[classes]["C1"] = "Student ID"
    formatted_workbook[classes]["D1"] = "Grade"

    # COMPLETED AS A GROUP - TASK 2 - In each new sheet, create separate columns for last name, first name, Student ID, and grade
    col_num = 2
    for row in range(2, old_workbook['Grades'].max_row + 1) :
        
        if classes == old_workbook['Grades'].cell(row=row, column=1).value :
            stud_string = old_workbook['Grades'].cell(row=row, column=2).value  # Full "Last_First_ID" string
            split_list = stud_string.split("_")
            # BLAKE ROGERS - add grade values to each row in each sheet
            grade = old_workbook['Grades'].cell(row=row, column=3).value

            split_list.append(grade)
            formatted_workbook[classes].append(split_list)

            
# HALEY SOMMER - TASK 3 - Each column should have an excel filter element above it

# Use the previously written class for loop as a basis for looping these columns in each sheet
    max_row = formatted_workbook[classes].max_row
    last_col = get_column_letter(formatted_workbook[classes].max_column) # use the openpxyl utility get_column_letter to fill in the last column letter
    formatted_workbook[classes].auto_filter.ref = f"A1:{last_col}{max_row}" # use auto_filter.ref to add filters to the first 4 columns

# Rebecca - TASK 4 - Each sheet should have summary information
# Use functions to calculate the following data
# The Highest Grade, lowest grade, mean grade, median grade, number of students in class
for classes in list_classes :
    formatted_workbook[classes]["F1"] = "Summary Type"
    formatted_workbook[classes]["F2"] = "Highest Grade"
    formatted_workbook[classes]["F3"] = "Lowest Grade"
    formatted_workbook[classes]["F4"] = "Median Grade"
    formatted_workbook[classes]["F5"] = "Number of Students"
    formatted_workbook[classes]["G1"] = "Data"
    formatted_workbook[classes]["G2"] = "=MAX(D:D)"
    formatted_workbook[classes]["G3"] = "=MIN(D:D)"
    formatted_workbook[classes]["G4"] = "=AVERAGE(D:D)"
    formatted_workbook[classes]["G5"] = "=COUNT(D:D)"


# ELISE CHAPMAN - TASK 5 - Format each sheet so the columns are the title width + 5.
# Bold Headers
columns = ["A", "B", "C", "D", "E", "F", "G"]
for classes in list_classes :
    i = 0
    for row in formatted_workbook[classes]["A1:G1"] :
        for cell in row:
            cell.font = Font(bold=True)
            max_width = len(str(cell.value))
            adjusted_width = max_width + 5
            formatted_workbook[classes].column_dimensions [columns[i]].width = adjusted_width
            i +=1

# CONRAD BRADFORD - TASK 6 - Save the new excel workbook named "formatted_grades.xlsx"

formatted_workbook.save(filename="formatted_grades.xlsx")
formatted_workbook.close()